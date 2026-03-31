"""
API endpoints para exportar cierres y comparaciones a CSV y Excel
"""
from fastapi import APIRouter, HTTPException, Depends, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from db.database import get_db
from db.models import CierreMensual, CierreMensualUsuario, Printer
from middleware.auth_middleware import get_current_user
from services.company_filter_service import CompanyFilterService

router = APIRouter(prefix="/api/export", tags=["export"])


def format_number_es(num: int) -> str:
    """Formatea número con separador de miles estilo español"""
    return f"{num:,}".replace(",", ".")


@router.get("/cierre/{cierre_id}", status_code=status.HTTP_200_OK)
async def export_cierre(
    cierre_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Exporta un cierre individual a CSV
    Formato: Usuario, Nombre, B/N, COLOR, TOTAL IMPRESIONES
    """
    cierre = db.query(CierreMensual).filter(CierreMensual.id == cierre_id).first()
    if not cierre:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cierre no encontrado")
    
    # Validar acceso a la impresora del cierre
    printer = db.query(Printer).filter(Printer.id == cierre.printer_id).first()
    if not printer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Impresora no encontrada")
    
    if not CompanyFilterService.validate_company_access(current_user, printer.empresa_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tienes acceso a este cierre")
    
    # Crear CSV en memoria
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Usuario', 'Nombre', 'B/N', 'COLOR', 'TOTAL IMPRESIONES', '', '', ''])
    
    # Usuarios ordenados por consumo descendente
    usuarios = sorted(cierre.usuarios, key=lambda u: u.consumo_total, reverse=True)
    
    suma_bn = 0
    suma_color = 0
    suma_total = 0
    
    for usuario in usuarios:
        # Solo incluir usuarios con consumo
        if usuario.consumo_total > 0:
            bn = usuario.impresora_bn + usuario.copiadora_bn + usuario.escaner_bn
            color = usuario.impresora_color + usuario.copiadora_color + usuario.escaner_color
            total = usuario.consumo_total
            
            writer.writerow([
                f'[{usuario.codigo_usuario}]',
                f'[{usuario.nombre_usuario}]',
                bn,
                color,
                total,
                '', '', ''
            ])
            
            suma_bn += bn
            suma_color += color
            suma_total += total
    
    # Fila con el total del contador de la impresora
    writer.writerow(['', '', '', '', '', '', '', cierre.total_paginas])
    
    # Fila final con suma de usuarios
    writer.writerow(['', '', '', '', suma_total, '', '', suma_total])
    
    # Preparar respuesta
    output.seek(0)
    # Formato: SERIAL DD.MM.YYYY
    fecha_actual = datetime.now().strftime('%d.%m.%Y')
    serial = printer.serial_number if printer and printer.serial_number else printer.hostname if printer else str(cierre.printer_id)
    filename = f"{serial} {fecha_actual}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/comparacion/{cierre1_id}/{cierre2_id}", status_code=status.HTTP_200_OK)
async def export_comparacion(
    cierre1_id: int,
    cierre2_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Exporta una comparación entre dos cierres a CSV
    Formato similar al comparativo manual: Usuario, Nombre, B/N, COLOR, TOTAL IMPRESIONES
    Con los totales de cada período al final
    """
    cierre1 = db.query(CierreMensual).filter(CierreMensual.id == cierre1_id).first()
    cierre2 = db.query(CierreMensual).filter(CierreMensual.id == cierre2_id).first()
    
    if not cierre1 or not cierre2:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uno o ambos cierres no encontrados")
    
    if cierre1.printer_id != cierre2.printer_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Los cierres deben ser de la misma impresora")
    
    # Validar acceso a la impresora
    printer = db.query(Printer).filter(Printer.id == cierre1.printer_id).first()
    if not printer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Impresora no encontrada")
    
    if not CompanyFilterService.validate_company_access(current_user, printer.empresa_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tienes acceso a esta impresora")
    
    # Crear mapa de usuarios
    usuarios_c1 = {u.codigo_usuario: u for u in cierre1.usuarios}
    usuarios_c2 = {u.codigo_usuario: u for u in cierre2.usuarios}
    
    # Todos los códigos únicos
    codigos = set(usuarios_c1.keys()).union(set(usuarios_c2.keys()))
    
    # Crear CSV en memoria
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Usuario', 'Nombre', 'B/N', 'COLOR', 'TOTAL IMPRESIONES', '', '', ''])
    
    # Calcular diferencias por usuario
    usuarios_diff = []
    suma_bn = 0
    suma_color = 0
    suma_total = 0
    
    for codigo in codigos:
        u1 = usuarios_c1.get(codigo)
        u2 = usuarios_c2.get(codigo)
        
        nombre = u2.nombre_usuario if u2 else (u1.nombre_usuario if u1 else codigo)
        
        # Calcular diferencia (puede ser negativa si hay correcciones)
        total_c1 = u1.total_paginas if u1 else 0
        total_c2 = u2.total_paginas if u2 else 0
        diff_total = total_c2 - total_c1
        
        # Calcular B/N y Color del período usando total_bn y total_color
        bn_c1 = u1.total_bn if u1 else 0
        bn_c2 = u2.total_bn if u2 else 0
        diff_bn = bn_c2 - bn_c1
        
        color_c1 = u1.total_color if u1 else 0
        color_c2 = u2.total_color if u2 else 0
        diff_color = color_c2 - color_c1
        
        # Incluir todos los usuarios (incluso con diferencia 0 o negativa)
        usuarios_diff.append({
            'codigo': codigo,
            'nombre': nombre,
            'bn': diff_bn,
            'color': diff_color,
            'total': diff_total
        })
        
        suma_bn += diff_bn
        suma_color += diff_color
        suma_total += diff_total
    
    # Ordenar por total descendente
    usuarios_diff.sort(key=lambda x: x['total'], reverse=True)
    
    # Escribir usuarios
    for u in usuarios_diff:
        writer.writerow([
            f'[{u["codigo"]}]',
            f'[{u["nombre"]}]',
            u['bn'],
            u['color'],
            u['total'],
            '', '', ''
        ])
    
    # Fila con total del período 1
    writer.writerow(['', '', '', '', '', '', '', cierre1.total_paginas])
    
    # Fila con total del período 2
    writer.writerow(['', '', '', '', '', '', '', cierre2.total_paginas])
    
    # Fila final con consumo y suma de usuarios
    diff_impresora = cierre2.total_paginas - cierre1.total_paginas
    writer.writerow(['', '', '', '', suma_total, '', '', diff_impresora])
    
    # Preparar respuesta
    output.seek(0)
    # Formato: SERIAL DD.MM.YYYY
    fecha_actual = datetime.now().strftime('%d.%m.%Y')
    serial = printer.serial_number if printer and printer.serial_number else printer.hostname if printer else str(cierre1.printer_id)
    filename = f"{serial} {fecha_actual}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )



@router.get("/cierre/{cierre_id}/excel", status_code=status.HTTP_200_OK)
async def export_cierre_excel(
    cierre_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Exporta un cierre individual a Excel
    Formato: Usuario, Nombre, B/N, COLOR, TOTAL IMPRESIONES
    """
    cierre = db.query(CierreMensual).filter(CierreMensual.id == cierre_id).first()
    if not cierre:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cierre no encontrado")
    
    # Validar acceso a la impresora del cierre
    printer = db.query(Printer).filter(Printer.id == cierre.printer_id).first()
    if not printer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Impresora no encontrada")
    
    if not CompanyFilterService.validate_company_access(current_user, printer.empresa_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tienes acceso a este cierre")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cierre"
    
    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Header
    headers = ['Usuario', 'Nombre', 'B/N', 'COLOR', 'TOTAL IMPRESIONES']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Usuarios ordenados por consumo descendente
    usuarios = sorted(cierre.usuarios, key=lambda u: u.consumo_total, reverse=True)
    
    suma_bn = 0
    suma_color = 0
    suma_total = 0
    
    row = 2
    for usuario in usuarios:
        # Solo incluir usuarios con consumo
        if usuario.consumo_total > 0:
            bn = usuario.impresora_bn + usuario.copiadora_bn + usuario.escaner_bn
            color = usuario.impresora_color + usuario.copiadora_color + usuario.escaner_color
            total = usuario.consumo_total
            
            ws.cell(row=row, column=1, value=f'[{usuario.codigo_usuario}]')
            ws.cell(row=row, column=2, value=f'[{usuario.nombre_usuario}]')
            ws.cell(row=row, column=3, value=bn)
            ws.cell(row=row, column=4, value=color)
            ws.cell(row=row, column=5, value=total)
            
            suma_bn += bn
            suma_color += color
            suma_total += total
            row += 1
    
    # Fila con el total del contador de la impresora
    ws.cell(row=row, column=8, value=cierre.total_paginas)
    ws.cell(row=row, column=8).font = Font(bold=True)
    row += 1
    
    # Fila final con suma de usuarios
    ws.cell(row=row, column=5, value=suma_total)
    ws.cell(row=row, column=5).font = Font(bold=True, color="0000FF")
    ws.cell(row=row, column=8, value=suma_total)
    ws.cell(row=row, column=8).font = Font(bold=True, color="0000FF")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Formato: SERIAL DD.MM.YYYY
    fecha_actual = datetime.now().strftime('%d.%m.%Y')
    serial = printer.serial_number if printer and printer.serial_number else printer.hostname if printer else str(cierre.printer_id)
    filename = f"{serial} {fecha_actual}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/comparacion/{cierre1_id}/{cierre2_id}/excel", status_code=status.HTTP_200_OK)
async def export_comparacion_excel(
    cierre1_id: int,
    cierre2_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Exporta una comparación entre dos cierres a Excel
    Formato: Muestra valores de ambos períodos + diferencias
    """
    cierre1 = db.query(CierreMensual).filter(CierreMensual.id == cierre1_id).first()
    cierre2 = db.query(CierreMensual).filter(CierreMensual.id == cierre2_id).first()
    
    if not cierre1 or not cierre2:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uno o ambos cierres no encontrados")
    
    if cierre1.printer_id != cierre2.printer_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Los cierres deben ser de la misma impresora")
    
    # Validar acceso a la impresora
    printer = db.query(Printer).filter(Printer.id == cierre1.printer_id).first()
    if not printer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Impresora no encontrada")
    
    if not CompanyFilterService.validate_company_access(current_user, printer.empresa_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tienes acceso a esta impresora")
    
    # Crear mapa de usuarios
    usuarios_c1 = {u.codigo_usuario: u for u in cierre1.usuarios}
    usuarios_c2 = {u.codigo_usuario: u for u in cierre2.usuarios}
    
    # Todos los códigos únicos
    codigos = set(usuarios_c1.keys()).union(set(usuarios_c2.keys()))
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Header simple con solo diferencias
    headers = ['Usuario', 'Nombre', 'B/N', 'COLOR', 'TOTAL IMPRESIONES']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Calcular diferencias por usuario
    usuarios_diff = []
    suma_bn = 0
    suma_color = 0
    suma_total = 0
    
    for codigo in codigos:
        u1 = usuarios_c1.get(codigo)
        u2 = usuarios_c2.get(codigo)
        
        nombre = u2.nombre_usuario if u2 else (u1.nombre_usuario if u1 else codigo)
        
        # Valores del período 1
        total_c1 = u1.total_paginas if u1 else 0
        bn_c1 = u1.total_bn if u1 else 0
        color_c1 = u1.total_color if u1 else 0
        
        # Valores del período 2
        total_c2 = u2.total_paginas if u2 else 0
        bn_c2 = u2.total_bn if u2 else 0
        color_c2 = u2.total_color if u2 else 0
        
        # Calcular diferencias (puede ser negativa si hay correcciones)
        diff_total = total_c2 - total_c1
        diff_bn = bn_c2 - bn_c1
        diff_color = color_c2 - color_c1
        
        # Incluir todos los usuarios (incluso con diferencia 0 o negativa)
        usuarios_diff.append({
            'codigo': codigo,
            'nombre': nombre,
            'bn': diff_bn,
            'color': diff_color,
            'total': diff_total
        })
        
        suma_bn += diff_bn
        suma_color += diff_color
        suma_total += diff_total
    
    # Ordenar por diferencia total descendente
    usuarios_diff.sort(key=lambda x: x['total'], reverse=True)
    
    # Escribir usuarios
    row = 2
    for u in usuarios_diff:
        ws.cell(row=row, column=1, value=f'[{u["codigo"]}]')
        ws.cell(row=row, column=2, value=f'[{u["nombre"]}]')
        ws.cell(row=row, column=3, value=u['bn'])
        ws.cell(row=row, column=4, value=u['color'])
        ws.cell(row=row, column=5, value=u['total'])
        row += 1
    
    # Fila de totales
    row += 1
    ws.cell(row=row, column=1, value="TOTALES")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=suma_bn)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value=suma_color)
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=suma_total)
    ws.cell(row=row, column=5).font = Font(bold=True, color="0000FF")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Formato: SERIAL DD.MM.YYYY
    fecha_actual = datetime.now().strftime('%d.%m.%Y')
    serial = printer.serial_number if printer and printer.serial_number else str(cierre1.printer_id)
    filename = f"{serial} {fecha_actual}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/comparacion/{cierre1_id}/{cierre2_id}/excel-ricoh", status_code=status.HTTP_200_OK)
async def export_comparacion_excel_ricoh(
    cierre1_id: int,
    cierre2_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Exporta una comparación entre dos cierres a Excel en formato Ricoh completo
    Formato: 3 hojas con 52 columnas cada una (Período 1, Período 2, Comparativo)
    Compatible con archivos originales de Ricoh
    """
    from services.export_ricoh import exportar_comparacion_ricoh
    
    cierre1 = db.query(CierreMensual).filter(CierreMensual.id == cierre1_id).first()
    cierre2 = db.query(CierreMensual).filter(CierreMensual.id == cierre2_id).first()
    
    if not cierre1 or not cierre2:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uno o ambos cierres no encontrados")
    
    if cierre1.printer_id != cierre2.printer_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Los cierres deben ser de la misma impresora")
    
    # Validar acceso a la impresora
    printer = db.query(Printer).filter(Printer.id == cierre1.printer_id).first()
    if not printer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Impresora no encontrada")
    
    if not CompanyFilterService.validate_company_access(current_user, printer.empresa_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tienes acceso a esta impresora")
    
    if not printer.serial_number:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La impresora debe tener un número de serie")
    
    # Generar archivo Excel en formato Ricoh
    wb = exportar_comparacion_ricoh(
        db=db,
        serial_number=printer.serial_number,
        cierre1=cierre1,
        cierre2=cierre2
    )
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombres de meses
    meses = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
             'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    
    mes1_nombre = meses[cierre1.mes]
    mes2_nombre = meses[cierre2.mes]
    
    # Formato: SERIAL DD.MM.YYYY
    fecha_actual = datetime.now().strftime('%d.%m.%Y')
    filename = f"{printer.serial_number} {fecha_actual}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
