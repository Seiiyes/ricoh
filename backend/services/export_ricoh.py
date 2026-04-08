"""
Servicio para exportar cierres en formato Excel de Ricoh (52 columnas)
Compatible con archivos originales de comparación
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from db.models import CierreMensual, CierreMensualUsuario

# Definición de columnas en orden exacto (52 columnas)
COLUMNAS_PERIODO = [
    'Usuario',
    'Nombre',
    'Total impresiones',
    'ByN(Total impresiones)',
    'Color(Total impresiones)',
    'ByN:Resultado(Total impresiones)',
    'Color:Resultado(Total impresiones)',
    'Blanco y negroTotal(Copiadora/Document Server)',
    'Blanco y negro(Tamaño pequeño)(Copiadora/Document Server)',
    'Blanco y negro(Tamaño grande)(Copiadora/Document Server)',
    'Mono ColorTotal(Copiadora/Document Server)',
    'Mono Color(Tamaño pequeño)(Copiadora/Document Server)',
    'Mono Color(Tamaño grande)(Copiadora/Document Server)',
    'Dos coloresTotal(Copiadora/Document Server)',
    'Dos colores(Tamaño pequeño)(Copiadora/Document Server)',
    'Dos colores(Tamaño grande)(Copiadora/Document Server)',
    'A Todo ColorTotal(Copiadora/Document Server)',
    'A Todo Color(Tamaño pequeño)(Copiadora/Document Server)',
    'A Todo Color(Tamaño grande)(Copiadora/Document Server)',
    'Blanco y negroTotal(Impresora)',
    'Blanco y negro(Tamaño pequeño)(Impresora)',
    'Blanco y negro(Tamaño grande)(Impresora)',
    'Mono ColorTotal(Impresora)',
    'Mono Color(Tamaño pequeño)(Impresora)',
    'Mono Color(Tamaño grande)(Impresora)',
    'Dos coloresTotal(Impresora)',
    'Dos colores(Tamaño pequeño)(Impresora)',
    'Dos colores(Tamaño grande)(Impresora)',
    'ColorTotal(Impresora)',
    'Color(Tamaño pequeño)(Impresora)',
    'Color(Tamaño grande)(Impresora)',
    'EscánerTotal(Escáner)',
    'Blanco y negroTotal(Escáner)',
    'Blanco y negro(Tamaño pequeño)(Escáner)',
    'Blanco y negro(Tamaño grande)(Escáner)',
    'A Todo ColorTotal(Escáner)',
    'A Todo Color(Tamaño pequeño)(Escáner)',
    'A Todo Color(Tamaño grande)(Escáner)',
    'Blanco y negroTotal(Fax)',
    'Blanco y negro(Tamaño pequeño)(Fax)',
    'Blanco y negro(Tamaño grande)(Fax)',
    'ColorTotal(Fax)',
    'Color(Tamaño pequeño)(Fax)',
    'Color(Tamaño grande)(Fax)',
    'Páginas transmitidas(Fax)',
    'Cargo por transmisión(Fax)',
    'Volumen usado(Limitación uso volumen impresión)',
    'Valor límite(Limitación uso volumen impresión)',
    'Volumen usado anterior(Limitación uso volumen impresión)',
    'Última fecha reinicio(Limitación uso volumen impresión)',
    'Negro(Revelado)',
    'Color (YMC)(Revelado)',
]


def formatear_codigo(codigo: str) -> str:
    """Formatea el código con corchetes"""
    return f"[{codigo}]"


def formatear_nombre(nombre: str) -> str:
    """Formatea el nombre con corchetes"""
    return f"[{nombre}]"


def crear_fila_usuario(usuario: CierreMensualUsuario, db: Session = None) -> list:
    """Crea una fila de datos para un usuario (52 columnas)"""
    # Obtener datos del usuario desde la tabla users
    from db.models import User
    user = db.query(User).filter(User.id == usuario.user_id).first() if db and usuario.user_id else None
    
    codigo = user.codigo_de_usuario if user else str(usuario.user_id)
    nombre = user.name if user else f"Usuario {usuario.user_id}"
    
    fila = [
        formatear_codigo(codigo),
        formatear_nombre(nombre),
        usuario.total_paginas,
        usuario.total_bn,
        usuario.total_color,
        usuario.total_bn,  # ByN:Resultado
        usuario.total_color,  # Color:Resultado
        usuario.copiadora_bn,
        0,  # Tamaño pequeño copiadora B/N
        0,  # Tamaño grande copiadora B/N
        0,  # Mono Color Total
        0,  # Mono Color pequeño
        0,  # Mono Color grande
        0,  # Dos colores Total
        0,  # Dos colores pequeño
        0,  # Dos colores grande
        usuario.copiadora_color,
        0,  # A Todo Color pequeño
        0,  # A Todo Color grande
        usuario.impresora_bn,
        0,  # Impresora B/N pequeño
        0,  # Impresora B/N grande
        0,  # Impresora Mono Color Total
        0,  # Impresora Mono Color pequeño
        0,  # Impresora Mono Color grande
        0,  # Impresora Dos colores Total
        0,  # Impresora Dos colores pequeño
        0,  # Impresora Dos colores grande
        usuario.impresora_color,
        0,  # Impresora Color pequeño
        0,  # Impresora Color grande
        usuario.escaner_bn + usuario.escaner_color,  # Escáner Total
        usuario.escaner_bn,
        0,  # Escáner B/N pequeño
        0,  # Escáner B/N grande
        usuario.escaner_color,
        0,  # Escáner Color pequeño
        0,  # Escáner Color grande
        usuario.fax_bn,
        0,  # Fax B/N pequeño
        0,  # Fax B/N grande
        0,  # Fax Color Total
        0,  # Fax Color pequeño
        0,  # Fax Color grande
        0,  # Páginas transmitidas
        0,  # Cargo por transmisión
        0,  # Volumen usado
        0,  # Valor límite
        0,  # Volumen usado anterior
        '',  # Última fecha reinicio
        usuario.total_bn,  # Negro (Revelado)
        usuario.total_color,  # Color (Revelado)
    ]
    
    return fila


def crear_fila_totales(usuarios: list) -> list:
    """Crea la fila de totales (52 columnas)"""
    totales = {
        'total_paginas': sum(u.total_paginas for u in usuarios),
        'total_bn': sum(u.total_bn for u in usuarios),
        'total_color': sum(u.total_color for u in usuarios),
        'copiadora_bn': sum(u.copiadora_bn for u in usuarios),
        'copiadora_color': sum(u.copiadora_color for u in usuarios),
        'impresora_bn': sum(u.impresora_bn for u in usuarios),
        'impresora_color': sum(u.impresora_color for u in usuarios),
        'escaner_bn': sum(u.escaner_bn for u in usuarios),
        'escaner_color': sum(u.escaner_color for u in usuarios),
        'fax_bn': sum(u.fax_bn for u in usuarios),
    }
    
    fila = [
        '',  # Usuario vacío
        '',  # Nombre vacío
        totales['total_paginas'],
        totales['total_bn'],
        totales['total_color'],
        totales['total_bn'],
        totales['total_color'],
        totales['copiadora_bn'],
        0, 0, 0, 0, 0, 0, 0, 0,
        totales['copiadora_color'],
        0, 0,
        totales['impresora_bn'],
        0, 0, 0, 0, 0, 0, 0, 0,
        totales['impresora_color'],
        0, 0,
        totales['escaner_bn'] + totales['escaner_color'],
        totales['escaner_bn'],
        0, 0,
        totales['escaner_color'],
        0, 0,
        totales['fax_bn'],
        0, 0, 0, 0, 0, 0, 0, 0, 0, 0, '',
        totales['total_bn'],
        totales['total_color'],
    ]
    
    return fila


def aplicar_formato_encabezado(ws):
    """Aplica formato a los encabezados"""
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(1, len(COLUMNAS_PERIODO) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def aplicar_formato_datos(ws, num_filas):
    """Aplica formato a los datos"""
    # Alinear números a la derecha
    for row in range(2, num_filas + 1):
        for col in range(3, len(COLUMNAS_PERIODO) + 1):  # Desde columna 3 (números)
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="right")
    
    # Negrita en última fila (totales)
    for col in range(1, len(COLUMNAS_PERIODO) + 1):
        cell = ws.cell(row=num_filas, column=col)
        cell.font = Font(bold=True)


def ajustar_anchos_columnas(ws):
    """Ajusta el ancho de las columnas"""
    ws.column_dimensions['A'].width = 12  # Usuario
    ws.column_dimensions['B'].width = 35  # Nombre
    
    # Resto de columnas
    for col in range(3, len(COLUMNAS_PERIODO) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def exportar_comparacion_ricoh(
    db: Session,
    serial_number: str,
    cierre1: CierreMensual,
    cierre2: CierreMensual
) -> Workbook:
    """
    Exporta una comparación de cierres en formato Excel de Ricoh
    3 hojas: Período 1 (completo), Período 2 (completo), Comparativo (solo diferencias)
    La hoja comparativa se adapta a las capacidades de la impresora
    
    Args:
        db: Sesión de base de datos
        serial_number: Serial de la impresora
        cierre1: Primer cierre (más antiguo - período base)
        cierre2: Segundo cierre (más reciente - período comparado)
    
    Returns:
        Workbook de openpyxl con 3 hojas
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from db.models import Printer
    
    # Obtener capacidades de la impresora
    printer = db.query(Printer).filter(Printer.id == cierre1.printer_id).first()
    has_color = printer.has_color if printer else False
    
    # Nombres de meses
    meses = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
             'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    
    mes1_nombre = meses[cierre1.mes]
    mes2_nombre = meses[cierre2.mes]
    
    # Crear workbook
    wb = Workbook()
    
    # ========== HOJA 1: Período 1 (completo con 52 columnas) ==========
    ws1 = wb.active
    ws1.title = f"{serial_number} {mes1_nombre}"
    
    # Encabezados
    ws1.append(COLUMNAS_PERIODO)
    aplicar_formato_encabezado(ws1)
    
    # Usuarios ordenados por user_id
    usuarios1 = sorted(cierre1.usuarios, key=lambda u: u.user_id)
    for usuario in usuarios1:
        fila = crear_fila_usuario(usuario, db)
        ws1.append(fila)
    
    # Fila de totales
    fila_totales1 = crear_fila_totales(usuarios1)
    ws1.append(fila_totales1)
    
    aplicar_formato_datos(ws1, ws1.max_row)
    ajustar_anchos_columnas(ws1)
    
    # ========== HOJA 2: Período 2 (completo con 52 columnas) ==========
    ws2 = wb.create_sheet(title=f"{serial_number} {mes2_nombre}")
    
    # Encabezados
    ws2.append(COLUMNAS_PERIODO)
    aplicar_formato_encabezado(ws2)
    
    # Usuarios ordenados por user_id
    usuarios2 = sorted(cierre2.usuarios, key=lambda u: u.user_id)
    for usuario in usuarios2:
        fila = crear_fila_usuario(usuario, db)
        ws2.append(fila)
    
    # Fila de totales
    fila_totales2 = crear_fila_totales(usuarios2)
    ws2.append(fila_totales2)
    
    aplicar_formato_datos(ws2, ws2.max_row)
    ajustar_anchos_columnas(ws2)
    
    # ========== HOJA 3: Comparativo (solo diferencias/consumo) ==========
    ws3 = wb.create_sheet(title=f"{serial_number} COMPARATIVO")
    
    # Encabezados adaptativos según capacidades de la impresora
    if has_color:
        # Impresora con color: mostrar B/N, Color y Total
        headers = ['Usuario', 'Nombre', 'B/N', 'COLOR', 'TOTAL IMPRESIONES', '', '']
        col_bn = 2  # Columna C (índice 2)
        col_color = 3  # Columna D (índice 3)
        col_total = 4  # Columna E (índice 4)
        col_periodo1 = 5  # Columna F (índice 5)
        col_periodo2 = 6  # Columna G (índice 6)
    else:
        # Impresora solo B/N: mostrar solo B/N (que es el total)
        headers = ['Usuario', 'Nombre', 'TOTAL IMPRESIONES', '', '', '', '']
        col_bn = 2  # Columna C (índice 2)
        col_color = None
        col_total = 2  # Columna C (índice 2) - mismo que B/N
        col_periodo1 = 5  # Columna F (índice 5)
        col_periodo2 = 6  # Columna G (índice 6)
    
    ws3.append(headers)
    
    # Estilo de encabezado
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # La fila de encabezados ahora es la 1
    header_row = 1
    for cell in ws3[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Crear diccionario de usuarios de ambos períodos por user_id
    usuarios1_dict = {u.user_id: u for u in cierre1.usuarios}
    usuarios2_dict = {u.user_id: u for u in cierre2.usuarios}
    
    # Obtener todos los user_ids únicos y ordenarlos
    todos_user_ids = sorted(set(usuarios1_dict.keys()) | set(usuarios2_dict.keys()))
    
    # Variables para sumar el consumo total de usuarios
    suma_consumo_bn = 0
    suma_consumo_color = 0
    suma_consumo_total = 0
    
    # Agregar filas de usuarios con DIFERENCIAS (consumo)
    for user_id in todos_user_ids:
        usuario1 = usuarios1_dict.get(user_id)
        usuario2 = usuarios2_dict.get(user_id)
        
        # Obtener datos del usuario desde la tabla users
        from db.models import User
        user = db.query(User).filter(User.id == user_id).first()
        codigo = user.codigo_de_usuario if user else str(user_id)
        nombre = user.name if user else f"Usuario {user_id}"
        
        # Calcular diferencias (consumo del período)
        total1 = usuario1.total_paginas if usuario1 else 0
        total2 = usuario2.total_paginas if usuario2 else 0
        consumo_total = total2 - total1
        
        bn1 = usuario1.total_bn if usuario1 else 0
        bn2 = usuario2.total_bn if usuario2 else 0
        consumo_bn = bn2 - bn1
        
        color1 = usuario1.total_color if usuario1 else 0
        color2 = usuario2.total_color if usuario2 else 0
        consumo_color = color2 - color1
        
        # Agregar fila según capacidades
        if has_color:
            ws3.append([
                formatear_codigo(codigo),
                formatear_nombre(nombre),
                consumo_bn,
                consumo_color,
                consumo_total,
                '',
                ''
            ])
            suma_consumo_bn += consumo_bn
            suma_consumo_color += consumo_color
        else:
            # Solo B/N (que es el total)
            ws3.append([
                formatear_codigo(codigo),
                formatear_nombre(nombre),
                consumo_total,
                '',
                '',
                '',
                ''
            ])
        
        suma_consumo_total += consumo_total
    
    # Calcular diferencia del contador total de la impresora
    total_cierre1 = cierre1.total_paginas
    total_cierre2 = cierre2.total_paginas
    diferencia_contador_total = total_cierre2 - total_cierre1
    
    # Fila vacía
    ws3.append(['', '', '', '', '', '', ''])
    
    # Fila con totales de cada período
    fila_periodos = ['', '', '', '', '', total_cierre1, total_cierre2]
    ws3.append(fila_periodos)
    
    # Fila con diferencias y validación
    diferencia_paginas_prueba = diferencia_contador_total - suma_consumo_total
    
    if has_color:
        # Con color: mostrar B/N, Color, Total, Diferencia contador, Páginas prueba
        fila_totales = ['', '', suma_consumo_bn, suma_consumo_color, suma_consumo_total, diferencia_contador_total, diferencia_paginas_prueba]
    else:
        # Solo B/N: mostrar Total, Diferencia contador, Páginas prueba
        fila_totales = ['', '', suma_consumo_total, diferencia_contador_total, diferencia_paginas_prueba, '', '']
    
    ws3.append(fila_totales)
    
    # Aplicar formato a las filas de totales
    last_row = ws3.max_row
    for cell in ws3[last_row]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    for cell in ws3[last_row - 1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Ajustar anchos de columna
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 18
    ws3.column_dimensions['G'].width = 18
    
    # Aplicar formato de números con separador de miles
    # Empezar desde la fila 2 (después del encabezado en fila 1)
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row[2:]:  # Columnas numéricas (C en adelante)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
    
    return wb
